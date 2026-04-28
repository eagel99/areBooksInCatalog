from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass
class Row:
    professor: str
    title: str
    author: str
    publisher: str
    source_row: int


@dataclass
class Result:
    professor: str
    title: str
    exists: bool
    permalink: str


def _cell(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def read_rows(path: str | Path) -> list[Row]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows: list[Row] = []
    for r in range(2, ws.max_row + 1):
        title = _cell(ws, r, 3)
        if not title:
            continue
        rows.append(
            Row(
                professor=_cell(ws, r, 1),
                title=title,
                author=_cell(ws, r, 4),
                publisher=_cell(ws, r, 7),
                source_row=r,
            )
        )
    wb.close()
    return rows


def write_results(path: str | Path, results: list[Result]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog Check"
    ws.sheet_view.rightToLeft = True

    headers = ["מי ביקש את הספר", "שם הספר", "קיים בקטלוג", "Permalink"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    link_font = Font(color="0563C1", underline="single")
    for i, res in enumerate(results, start=2):
        ws.cell(row=i, column=1, value=res.professor)
        ws.cell(row=i, column=2, value=res.title)
        ws.cell(row=i, column=3, value="כן" if res.exists else "לא")
        link_cell = ws.cell(row=i, column=4, value=res.permalink or "")
        if res.permalink:
            link_cell.hyperlink = res.permalink
            link_cell.font = link_font

    widths = [22, 60, 14, 80]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.freeze_panes = "A2"

    wb.save(path)
